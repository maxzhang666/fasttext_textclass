# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

import fasttext
from tqdm import tqdm
# from lib.cut_sentence
from openpyxl import load_workbook
import sacremoses
import sentencepiece as spm

fasttext.FastText.eprint = lambda x: None
# 标签位置  1/前   2/后
label_postion = 2
# bpe模型
bpe_model = spm.SentencePieceProcessor()
bpe_model.Load("./bpe_en.model")
# 标签映射
label_dic = {
    '质量-其他': 'Quality_Other',
    '质量-破损': 'Quality_Broken',
    '投递-异常': 'Delivery_Exception',
    '质量-少配件': 'Quality_Few_Ingredients',
    '投递-咨询': 'Delivery_Consultation',
    '投递-变更': 'Delivery_Change',
    '订单-订单取消': 'Order_Order_Cancellation',
    '发票-发票': 'Invoice',
    '订单-订单变更': 'Order_Order_Changes',
    '投递-异常、退款': 'Delivery_Abnormal',
    '发票-发票变更': 'Invoice_Invoice_Change',
    '质量-产品破损': 'Quality_Product_Breakage',
    '其他': 'Other'
}


def token_normalizer(data_path, out_path):
    nor_en = sacremoses.MosesPunctNormalizer(lang='en')
    tok_en = sacremoses.MosesTokenizer(lang='en')

    writer = open(out_path, "a")

    for line in open(data_path).readlines():
        en_nor = nor_en.normalize(line)
        en_tok = ' '.join(tok_en.tokenize(en_nor))
        writer.write(en_tok)


def create_ru_subword_dict(data_path):
    spm.SentencePieceTrainer.Train(input=data_path, model_prefix='sub_word_ru', vocab_size=32000)


def sub_word(line):
    encoder = bpe_model
    tokens = []
    for token, ids in list(zip(encoder.Encode(line.strip(), out_type=str), encoder.Encode(line.strip(), out_type=int))):
        if encoder.IsUnknown(ids):
            tokens += list(token)
        else:
            tokens.append(token)
    return ' '.join(tokens)


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


def process_excel_data(file, input_data, bpe=False):
    excel = load_workbook(input_data)
    table = excel[excel.sheetnames[0]]
    for i in table.rows:
        label = label_dic.get(i[1].value)
        data = i[0].value.replace("\n", " ")
        if bpe:
            data = sub_word(data)
        if label_postion == 1:
            file.write("__label__{} \t {} \n".format(label, data))
        else:
            file.write("{} \t__label__{}\n".format(data, label))


def model_train_old(model_path, data_path):
    model = fasttext.load_model(model_path)
    model = model.train_supervised(input=data_path, epoch=25, lr=0.6, wordNgrams=2, verbose=2, minCount=1)
    model.save_model(model_path)


def train(model_path, data_path, auto_arg=False, validation_file='', auto_time=600):
    if auto_arg:
        new_model = fasttext.train_supervised(input=data_path, autotuneValidationFile=validation_file,
                                              autotuneDuration=auto_time)
    else:
        new_model = fasttext.train_supervised(input=data_path, epoch=25, lr=0.6, wordNgrams=4, verbose=2, minCount=1)
    new_model.save_model(model_path)


def model_test(model_path, data_path):
    model = fasttext.load_model(model_path)
    res = model.test(data_path)
    print(res)


def model_red(model_path, data_path):
    model = fasttext.load_model(model_path)
    print(model.words)
    list = []
    for line in open(data_path).readlines():
        list.append(line.replace("\n", ""))
    res = model.predict(list)

    for i in res:
        print(i)

    # Press the green button in the gutter to run the script.


if __name__ == '__main__':
    # 训练集  普通训练集，BPE训练集
    # 测试集  普通测试集，BPE测试集

    # 定参训练      普通训练 BPE参数
    # 自动超参训练   普通训练 BPE训练 20分钟

    # 测试 定参 普通、BPE
    # 测试 自动 普通、BPE

    # 数据整理 测试集
    # process_excel_data(open('./test', "a"), "./test.xlsx")

    # 数据整理 PBE数据集
    # process_excel_data(open('./data_bpe', "w"), "./train.xlsx", True)
    # 数据整理 PBE测试集
    # process_excel_data(open('./test_bpe', "w"), "./test.xlsx", True)

    # 模型训练
    # train(r"./m1.bin", "./data")
    # 模型训练 bpe  手动参数
    # train(r"./m1_bpe_m.bin", "./data_bpe")
    # 数据训练 bpe  自动搜索
    # train(r"./m1_pbe.bin", "./data_bpe", True, "./test_bpe")
    # 模型测试
    # model_test("./m1.bin", "./test")
    # 模型测试 bpe
    # model_test("./m1_bpe.bin", "./test_bpe")
    # 模型测试 bpe 手工参数
    # model_test("./m1_bpe_m.bin", "./test_bpe")
    # 模型预测
    # model_red("./m1.bin", "./test")

    model = fasttext.load_model("./m1.bin")
    print(model.words)
